# -*- coding: utf-8 -*-
"""
Created on Tue Jul  8 15:54:20 2025

@author: Xuan Tung Duong
email: duongxt@student.ubc.ca

Last updated: Jul 26, 2025
"""
# import required libraries
import os
os.environ["OMP_NUM_THREADS"] = "1" # necessary environment change to run k-means
import cv2
import numpy as np
import pandas as pd
from datetime import datetime, timedelta
import matplotlib
matplotlib.use('Qt5Agg')  # force interactive backend for manual droplet detection
import matplotlib.pyplot as plt
from scipy.signal import find_peaks
from openpyxl import Workbook, load_workbook
from sklearn.cluster import KMeans

plt.ion()  # enable interactive plotting for Spyder

# === PARAMETERS ===
BRIGHTNESS_PEAK_PROMINENCE = 30  # brightness derivative threshold for freezing detection (1/s)
# lower this if droplets are detected but no freezing point is determined
MIN_RADIUS = 6  # min radius for droplet detection (pixels)
# change this if droplets aren't being detected properly
BINARY_THRESHOLD = 139  # brightness threshold for binary conversion (unitless)
# change this if there is a lot of noise in the droplet detection or if droplets aren't detected

# === SET WORKING DIRECTORY TO SCRIPT LOCATION ===
os.chdir(os.path.dirname(os.path.abspath(__file__)))

# prepare arrays to record data
all_results_by_folder = {}
all_results_by_plate = {}

# === DEFINE FUNCTIONS ===
def temp_at_time(tsv_df, sec, video_start_time): # TSV to Video temperature alignment
    video_abs_time = video_start_time + timedelta(seconds=int(sec))
    idx = (tsv_df["Timestamp"] - video_abs_time).abs().idxmin()
    return tsv_df.loc[idx, "Temperature"]

def prompt_manual_droplet_selection(image, detected_droplets): # manual droplet detection function
    selected = detected_droplets.copy()
    fig, ax = plt.subplots()
    ax.imshow(cv2.cvtColor(image, cv2.COLOR_BGR2RGB))
    circles = []

    for i, (cx, cy, r) in enumerate(selected):
        circ = plt.Circle((cx, cy), r, color='b', fill=False)
        circles.append(circ)
        ax.add_patch(circ)
        ax.text(cx + 5, cy - 5, str(i + 1), color='b', fontsize=8)

    def on_click(event):
        nonlocal selected, circles
        if event.inaxes != ax:
            return

        if event.button == 1:  # Left-click to add
            selected.append((int(event.xdata), int(event.ydata), MIN_RADIUS))
            redraw()
        elif event.button == 3:  # Right-click to remove
            for i, (cx, cy, r) in enumerate(selected):
                if np.hypot(cx - event.xdata, cy - event.ydata) <= r:
                    selected.pop(i)
                    redraw()
                    return

    def redraw():
        ax.clear()
        ax.imshow(cv2.cvtColor(image, cv2.COLOR_BGR2RGB))
        for cx, cy, r in selected:
            circ = plt.Circle((cx, cy), r, color='b', fill=False)
            ax.add_patch(circ)
            ax.text(cx + 5, cy - 5, str(selected.index((cx, cy, r)) + 1), color='b', fontsize=8)
        fig.canvas.draw()

    cid = fig.canvas.mpl_connect('button_press_event', on_click)
    plt.title("Left-click to add, right-click to remove droplets. Close window when done.")
    plt.show(block=True)

    fig.canvas.mpl_disconnect(cid)
    return selected

# === WALK THROUGH ALL SUBFOLDERS ===
base_dir = os.getcwd()

today = datetime.now().strftime("%Y_%m_%d")
output_filename = f"{today}_freezing_results.xlsx"
wb = Workbook()
if os.path.exists(output_filename):
    wb = load_workbook(output_filename)
    existing_sheets = wb.sheetnames
else:
    wb.remove(wb.active)

for folder_name in os.listdir(base_dir): # cycle through every subfolder in the same folder as the code
    root = os.path.join(base_dir, folder_name)
    if not os.path.isdir(root):
        continue

    print(f"\n--- Detected folder: {folder_name} ---")
    answer = input(f"Analyze folder '{folder_name}'? (Y/N): ").strip().upper()
    if answer != 'Y':
        continue

    files = os.listdir(root)
    video_files = [f for f in files if f.lower().endswith(('.wmv', '.mov', '.mp4'))] # add video file extensions here if needed
    tsvs = [f for f in files if f.lower().endswith('.tsv')]

    if len(video_files) != 1 or len(tsvs) != 1: 
        print(f"Skipping folder '{folder_name}' — requires exactly one video file and one .tsv")
        continue

    video_path = os.path.join(root, video_files[0])
    tsv_path = os.path.join(root, tsvs[0])
    sample_name = folder_name

    print(f"Video: {video_files[0]}")
    print(f"TSV: {tsvs[0]}")
    optimal_k = int(input("\nEnter number of plates: "))

    # === DROPLET DETECTION ===
    # converts final frame to binary image
    # uses contours to detect circles (white) against background (black)
    cap = cv2.VideoCapture(video_path)
    fps = cap.get(cv2.CAP_PROP_FPS)
    frame_count = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
    duration = int(frame_count / fps)

    frame = None
    for i in range(1, 11):
        cap.set(cv2.CAP_PROP_POS_FRAMES, frame_count - i)
        ret, candidate = cap.read()
        if ret and candidate is not None:
            frame = candidate
            break
    if frame is None:
        print(f"Could not read final frame for {sample_name}")
        continue

    hsv = cv2.cvtColor(frame, cv2.COLOR_BGR2HSV)
    h, s, v = cv2.split(hsv)
    s = np.clip(s * 1.3, 0, 255).astype(np.uint8)
    hsv = cv2.merge([h, s, v])
    frame = cv2.cvtColor(hsv, cv2.COLOR_HSV2BGR)
    
    frame = cv2.cvtColor(hsv.astype(np.uint8), cv2.COLOR_HSV2BGR)
    h, w = frame.shape[:2]
    x1, x2 = int(w * 0.2), int(w * 0.8)
    frame = frame[:, x1:x2]
    gray_init = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
    blurred = cv2.GaussianBlur(gray_init, (5, 5), 0)

    _, binary_image = cv2.threshold(blurred, BINARY_THRESHOLD, 255, cv2.THRESH_BINARY)
    binary_vis = cv2.cvtColor(binary_image, cv2.COLOR_GRAY2BGR)

    contours, _ = cv2.findContours(binary_image, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    droplets = []
    for cnt in contours:
        area = cv2.contourArea(cnt)
        if area < 10 or area > 1000:
            continue
        perimeter = cv2.arcLength(cnt, True)
        if perimeter == 0:
            continue
        circularity = 4 * np.pi * area / (perimeter ** 2)
        if 0.6 < circularity < 1.3:
            M = cv2.moments(cnt)
            if M["m00"] != 0:
                cx = int(M["m10"] / M["m00"])
                cy = int(M["m01"] / M["m00"])
                (x, y), radius = cv2.minEnclosingCircle(cnt)
                radius = int(radius * 1.2)
                if radius >= MIN_RADIUS:
                    droplets.append((cx, cy, radius))

    droplets = sorted(droplets, key=lambda d: d[0])
    print(f"\nDetected {len(droplets)} droplets in {folder_name}")

    manual = input(f"Manual droplet detection for {folder_name}? (Y/N): ").strip().upper()
    if manual == 'Y':
        droplets = prompt_manual_droplet_selection(frame, droplets)
        print(f"Droplet count after manual correction: {len(droplets)}")

    if not droplets:
        continue

    # save binary image with circles
    binary_output = binary_vis.copy()
    for i, (cx, cy, r) in enumerate(droplets, start=1):
        cv2.circle(binary_output, (cx, cy), r, (0, 255, 0), 2)
        cv2.putText(binary_output, str(i), (cx + 5, cy - 5), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 255, 0), 1)
    cv2.imwrite(os.path.join(root, "binary_detected_droplets.png"), binary_output)

    # save original image with circles
    output = frame.copy()    
    droplet_coords = np.array([(cx, cy) for cx, cy, _ in droplets])

    # cluster droplets using kmean to identify plates
    kmeans = KMeans(n_clusters=optimal_k, random_state=0).fit(droplet_coords)
    plate_labels = kmeans.labels_
    
    # annotate the plate circle once per plate
    output_frame_plate = output.copy()
    for i, (cx, cy, r) in enumerate(droplets, start=1):
        cv2.circle(output_frame_plate, (cx, cy), r, (255, 0, 0), 2)
        cv2.putText(output_frame_plate, str(i), (cx + 5, cy - 5), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 0, 0), 1)
    for plate_id in range(optimal_k):
        cluster_points = droplet_coords[plate_labels == plate_id]
        if len(cluster_points) == 0:
            continue
        x_mean, y_mean = np.mean(cluster_points, axis=0).astype(int)
        radius = int(np.max(np.linalg.norm(cluster_points - [x_mean, y_mean], axis=1))*1.1 )
        cv2.circle(output_frame_plate, (x_mean, y_mean), radius, (0, 255, 0), 2)
        cv2.putText(output_frame_plate, f"Plate {plate_id+1}", (x_mean - 30, y_mean - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0), 1)
    
    # save plate plot
    plt.figure(figsize=(12, 6))
    plt.imshow(cv2.cvtColor(output_frame_plate, cv2.COLOR_BGR2RGB))
    plt.title("Detected Droplets and Plates")
    plt.axis('off')
    plt.tight_layout()
    plt.savefig(os.path.join(root, "detected_droplets_and_plates.png"), dpi=300)
    plt.close()

    # prompt user for time parameters
    video_start_str = input("\nEnter video start time (HH:MM:SS, 24h format): ")
    file_start_str = input("Enter file (TSV) start time (HH:MM:SS, 24h format): ")
    skip_seconds = int(input("Enter number of skipped seconds: "))

    print("Processing video...")

    video_start_time = datetime.strptime(video_start_str, "%H:%M:%S")
    file_start_time = datetime.strptime(file_start_str, "%H:%M:%S")
    time_diff_sec = int((video_start_time - file_start_time).total_seconds())

    # prepare datafile for recording data
    df = pd.read_csv(tsv_path, sep='\t', skiprows=12, header=None, usecols=[0, 2], names=["Elapsed", "Temperature"])
    df["Elapsed"] = pd.to_datetime(df["Elapsed"], format="%H:%M:%S", errors='coerce')
    df = df.dropna(subset=["Elapsed"])
    df["Timestamp"] = df["Elapsed"].apply(lambda t: file_start_time + timedelta(hours=int(t.hour), minutes=int(t.minute), seconds=int(t.second)))

    droplet_brightness = {i + 1: [] for i in range(len(droplets))}

    # get brightness value for every frame
    cap.set(cv2.CAP_PROP_POS_FRAMES, 0)
    frame_idx = 0
    while True:
        ret, frame = cap.read()
        if not ret:
            break
    
        if frame_idx < skip_seconds * fps:
            frame_idx += 1
            continue
    
        frame = frame[:, x1:x2]
        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
    
        for i, (cx, cy, r) in enumerate(droplets):
            mask = np.zeros_like(gray, dtype=np.uint8)
            cv2.circle(mask, (cx, cy), r, 255, -1)
            mean_val = cv2.mean(gray, mask=mask)[0]
            droplet_brightness[i + 1].append(mean_val)
    
        frame_idx += 1

    cap.release()

    # compute brightness derivative and capture peaks for all droplets
    peak_info = []
    droplet_to_plate = {i + 1: plate_labels[i] + 1 for i in range(len(plate_labels))}  # Map droplet ID to plate ID
    plate_data = {pid: [] for pid in range(1, optimal_k + 1)}  # Prepare per-plate storage
    plt.figure(figsize=(12, 6))
    for droplet_id, brightness in droplet_brightness.items():
        brightness = np.array(brightness)
        deriv = np.diff(brightness) * fps # convert brightness deriv to 1/s
        frame_num = np.arange(len(deriv)) # index begins at 0
        sec_idx = (frame_num + 1) / fps + skip_seconds # +1 to correct frame_num index
        peaks, _ = find_peaks(deriv, prominence=BRIGHTNESS_PEAK_PROMINENCE) # returns frame num of deriv peaks
        
        max_peak = np.max(deriv[peaks]) # find highest brightness derivative
        max_peak_idx = np.argmax(deriv[peaks]) # index highest peak
        max_peak_frame = peaks[max_peak_idx] # frame num of highest peak (offset)
        peak_sec = (max_peak_frame + 1) / fps + skip_seconds # +1 to correct index

        # plot data onto brightness derivative graph
        plt.plot(sec_idx, deriv, label = f"Droplet {droplet_id}")
        plt.text(peak_sec, max_peak, str(droplet_id), fontsize=8, color="k")

        if len(peaks) < 1 or np.max(deriv[peaks]) < BRIGHTNESS_PEAK_PROMINENCE: # check if droplet freezes
            peak_info.append((droplet_id, None, None))  # Droplet didn't freeze (or under threshold) # append blanks       
        else: # append droplet data
            temp = temp_at_time(df, round(peak_sec), video_start_time)
            peak_info.append((droplet_id, peak_sec + time_diff_sec, temp))

    # generate brightness derivative graph
    plt.xlim(left = skip_seconds)
    plt.xlabel("Video Time (s)")
    plt.ylabel("First Derivative of Brightness ($s^{-1}$)")
    plt.title(f"{sample_name}: First Derivative of Brightness Plot")
    plt.grid(True)
    plt.tight_layout()
    plt.savefig(os.path.join(root, "brightness_derivative_plot.png"), dpi=300)
    plt.close()
    
    # prepare datafile for excel output
    result_df = pd.DataFrame(peak_info, columns=["Droplet #", "Freezing Time (s)", "Freezing Temp (C)"])

    # sort by freezing temp
    result_df = result_df.sort_values(by="Freezing Temp (C)", ascending=False, na_position='first').reset_index(drop=True)
    result_df["Frozen #"] = result_df["Freezing Temp (C)"].notna().cumsum().astype("Int64")
    result_df["Plate"] = result_df["Droplet #"].map(droplet_to_plate)
    all_results_by_folder[folder_name + "_by_temp"] = result_df.copy()
    
    # sort by plate number
    result_sorted_by_plate = result_df.sort_values(by=["Plate", "Droplet #"]).reset_index(drop=True)
    all_results_by_folder[folder_name + "_by_plate"] = result_sorted_by_plate
    
    print("Analysis complete!")

# === EXPORT ALL RESULTS ===
if all_results_by_folder:
    today_str = datetime.today().strftime("%Y_%m_%d")
    excel_filename = f"{today_str}_freezing_results.xlsx"
    writer = pd.ExcelWriter(excel_filename, engine='openpyxl')
    for folder, df_sheet in all_results_by_folder.items():
        sheet_name = folder[:31]
        df_sheet.index = range(1, len(df_sheet) + 1)
        df_sheet.to_excel(writer, sheet_name=sheet_name, startrow=2)
        ws = writer.sheets[sheet_name]
        ws.cell(row=1, column=1).value = f"Folder: {folder}"
        ws.cell(row=2, column=1).value = f"Date: {datetime.today().strftime('%Y-%m-%d')}"
    writer.close()
    print(f"\nSaved: {excel_filename}")
else:
    print("No results found across all folders.")

